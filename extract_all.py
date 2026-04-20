# -*- coding: utf-8 -*-
import sys, io, os, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

def parse_coord(val):
    if val is None: return None
    try:
        s = str(val).strip().replace(',', '').replace(' ', '')
        f = float(s)
        return f
    except:
        return None

def is_valid_coord(lat, lng):
    return (lat is not None and lng is not None and
            -90 <= lat <= 90 and -180 <= lng <= 180 and
            lat > 0 and lng > 100)

def extract_excel(fpath, kv_label):
    """Extract all sheets from an Excel file, return {sheet_name: [rows]}"""
    wb = openpyxl.load_workbook(fpath, data_only=True)
    print(f'\nFile: {os.path.basename(fpath)}')
    print(f'Sheets: {wb.sheetnames}')

    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Print first few rows to understand structure
        print(f'\n  Sheet: {sheet_name}  max_row={ws.max_row}')
        for r in range(1, 7):
            row_vals = [ws.cell(r, c).value for c in range(1, 12)]
            print(f'    Row {r}: {row_vals}')

        rows = []
        for r in range(6, ws.max_row + 1):
            tuyen   = ws.cell(r, 1).value
            stt     = ws.cell(r, 2).value
            don_vi  = ws.cell(r, 3).value
            olt     = ws.cell(r, 4).value
            diem_a  = ws.cell(r, 5).value
            diem_b  = ws.cell(r, 6).value
            lon_a   = parse_coord(ws.cell(r, 7).value)
            lat_a   = parse_coord(ws.cell(r, 8).value)
            lon_b   = parse_coord(ws.cell(r, 9).value)
            lat_b   = parse_coord(ws.cell(r, 10).value)
            cap_6fo = ws.cell(r, 11).value
            cap_12fo= ws.cell(r, 12).value
            cap_24fo= ws.cell(r, 13).value
            sp      = ws.cell(r, 14).value

            if stt is None and lon_a is None: continue
            if not is_valid_coord(lat_a, lon_a) and not is_valid_coord(lat_b, lon_b): continue

            rows.append({
                'tuyen':   str(tuyen)  if tuyen  else '',
                'stt':     stt,
                'don_vi':  str(don_vi) if don_vi  else '',
                'olt':     str(olt)    if olt     else '',
                'diem_a':  str(diem_a) if diem_a  else '',
                'diem_b':  str(diem_b) if diem_b  else '',
                'lon_a':   lon_a,
                'lat_a':   lat_a,
                'lon_b':   lon_b,
                'lat_b':   lat_b,
                'cap_6fo': cap_6fo,
                'cap_12fo':cap_12fo,
                'cap_24fo':cap_24fo,
                'sp':      sp,
                'kv':      kv_label,
            })
        result[sheet_name] = rows
        print(f'  => {len(rows)} valid rows')
    return result

# Find files
files = os.listdir('.')
kv5_file = [f for f in files if 'KV5' in f and f.endswith('.xlsx') and not f.startswith('~')][0]
kv6_file = [f for f in files if 'KV6' in f and f.endswith('.xlsx') and not f.startswith('~')][0]

print(f'KV5: {kv5_file}')
print(f'KV6: {kv6_file}')

kv5_data = extract_excel(kv5_file, 'KV5')
kv6_data = extract_excel(kv6_file, 'KV6')

# Merge into one JSON with kv prefix
all_data = {}
for k, v in kv5_data.items():
    all_data[f'KV5__{k}'] = v
for k, v in kv6_data.items():
    all_data[f'KV6__{k}'] = v

with open('map_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print('\n\nSummary:')
for k, v in all_data.items():
    print(f'  {k}: {len(v)} rows')
