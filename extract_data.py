# -*- coding: utf-8 -*-
import sys
import io
import os
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

files = os.listdir('.')
excel_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('~')]
fname = excel_files[0]
print(f"Loading: {fname}")
wb = openpyxl.load_workbook(fname, data_only=True)
print('Sheets:', wb.sheetnames)

all_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_rows = []
    
    # Data starts from row 6
    for r in range(6, ws.max_row + 1):
        tuyen = ws.cell(r, 1).value   # A - TUYẾN TỔNG HỢP
        stt = ws.cell(r, 2).value     # B - STT
        don_vi = ws.cell(r, 3).value  # C - Đơn vị
        olt = ws.cell(r, 4).value     # D - OLT
        diem_a = ws.cell(r, 5).value  # E - ĐIỂM ĐẦU A
        diem_b = ws.cell(r, 6).value  # F - ĐIỂM ĐẦU B
        lon_a = ws.cell(r, 7).value   # G - Long A
        lat_a = ws.cell(r, 8).value   # H - Lat A
        lon_b = ws.cell(r, 9).value   # I - Long B
        lat_b = ws.cell(r, 10).value  # J - Lat B
        cap_6fo = ws.cell(r, 11).value  # K
        cap_12fo = ws.cell(r, 12).value # L
        cap_24fo = ws.cell(r, 13).value # M
        sp = ws.cell(r, 14).value       # N
        
        # Skip empty rows (no STT or no coordinates)
        if stt is None and lon_a is None:
            continue
        
        # Try to parse coordinates
        def parse_coord(val):
            if val is None:
                return None
            try:
                # Handle string with spaces or commas
                s = str(val).strip().replace(',', '').replace(' ', '')
                return float(s)
            except:
                return None
        
        lon_a_f = parse_coord(lon_a)
        lat_a_f = parse_coord(lat_a)
        lon_b_f = parse_coord(lon_b)
        lat_b_f = parse_coord(lat_b)
        
        if lon_a_f is None and lat_a_f is None and lon_b_f is None and lat_b_f is None:
            continue
        
        row_data = {
            'tuyen': str(tuyen) if tuyen else '',
            'stt': stt,
            'don_vi': str(don_vi) if don_vi else '',
            'olt': str(olt) if olt else '',
            'diem_a': str(diem_a) if diem_a else '',
            'diem_b': str(diem_b) if diem_b else '',
            'lon_a': lon_a_f,
            'lat_a': lat_a_f,
            'lon_b': lon_b_f,
            'lat_b': lat_b_f,
            'cap_6fo': cap_6fo,
            'cap_12fo': cap_12fo,
            'cap_24fo': cap_24fo,
            'sp': sp
        }
        sheet_rows.append(row_data)
    
    all_data[sheet_name] = sheet_rows
    print(f"Sheet {sheet_name}: {len(sheet_rows)} valid rows")

# Save to JSON
with open('map_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nTotal data saved to map_data.json")
for sheet, rows in all_data.items():
    print(f"  {sheet}: {len(rows)} rows")
