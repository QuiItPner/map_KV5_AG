# -*- coding: utf-8 -*-
import sys
import io
import os
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

# Find the excel file
files = os.listdir('.')
excel_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('~')]
print("Excel files found:", excel_files)

if excel_files:
    fname = excel_files[0]
    print(f"Loading: {fname}")
    wb = openpyxl.load_workbook(fname)
    print('Sheets:', wb.sheetnames)
    
    all_data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n--- Sheet: {sheet_name} ---')
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
        
        # Print header row (row 1)
        header = []
        for c in range(1, min(ws.max_column+1, 15)):
            cell_val = ws.cell(1, c).value
            header.append(cell_val)
        print('Headers:', header)
        
        # Print first 5 data rows
        for r in range(2, min(7, ws.max_row+1)):
            row_data = []
            for c in range(1, min(ws.max_column+1, 15)):
                row_data.append(ws.cell(r, c).value)
            print(f'Row {r}:', row_data)
        
        # Specifically check G,H,I,J columns (7,8,9,10)
        print(f'\nColumns G,H,I,J data:')
        for r in range(1, min(6, ws.max_row+1)):
            g = ws.cell(r, 7).value
            h = ws.cell(r, 8).value
            i = ws.cell(r, 9).value
            j = ws.cell(r, 10).value
            print(f'Row {r}: G={g}, H={h}, I={i}, J={j}')
